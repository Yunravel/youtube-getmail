from __future__ import annotations

import json
import logging
import re
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from urllib.parse import parse_qs, unquote, urlparse

import requests

from .contact_parser import extract_public_contacts
from .crawler import BrowserCrawler, classify_email_status
from .email_finder import PublicEmailFinder, SOCIAL_HOSTS


SUPPORTED_PLATFORMS = {"youtube", "instagram", "tiktok", "x", "twitter"}
RESULT_COLUMNS = (
    "联系邮箱",
    "邮箱状态",
    "邮箱来源",
    "公开外链",
    "采集状态",
    "采集时间",
)
PUBLIC_URL_RE = re.compile(r"https?://[^\s<>\"']+", re.I)


@dataclass(frozen=True)
class ProfileResult:
    email: str = ""
    email_status: str = "未发现"
    email_source: str = ""
    public_links: str = ""
    crawl_status: str = "成功"


def normalize_platform(value: object) -> str:
    platform = str(value or "").strip().casefold()
    return "x" if platform == "twitter" else platform


def profile_url(platform: str, account: object, raw_url: object) -> str:
    url = str(raw_url or "").strip()
    if url:
        return url if "://" in url else "https://" + url
    name = str(account or "").strip().lstrip("@")
    if not name:
        return ""
    if platform == "youtube":
        if name.startswith("UC") and len(name) >= 20:
            return f"https://www.youtube.com/channel/{name}"
        return f"https://www.youtube.com/@{name}"
    if platform == "instagram":
        return f"https://www.instagram.com/{name}/"
    if platform == "tiktok":
        return f"https://www.tiktok.com/@{name}"
    if platform == "x":
        return f"https://x.com/{name}"
    return ""


def unwrap_external_url(url: str) -> str:
    """Unwrap common social redirect links without making a request."""
    parsed = urlparse(url)
    host = (parsed.hostname or "").casefold()
    query = parse_qs(parsed.query)
    if host in {"l.instagram.com", "lm.instagram.com"} and query.get("u"):
        return unquote(query["u"][0])
    if host.endswith("youtube.com") and parsed.path == "/redirect" and query.get("q"):
        return unquote(query["q"][0])
    return url


def select_public_links(profile_url_value: str, links: list[str]) -> list[str]:
    selected: list[str] = []
    profile_host = (urlparse(profile_url_value).hostname or "").casefold()
    for raw in links:
        url = unwrap_external_url(str(raw or "").strip())
        if not url or url.startswith(("javascript:", "#")):
            continue
        if url.startswith("mailto:"):
            if url not in selected:
                selected.append(url)
            continue
        parsed = urlparse(url)
        host = (parsed.hostname or "").casefold()
        if parsed.scheme not in {"http", "https"} or not host:
            continue
        # Keep only links deliberately leading away from the current social site.
        if host == profile_host or host in SOCIAL_HOSTS:
            continue
        if url not in selected:
            selected.append(url)
    return selected


def extract_tiktok_profile_data(documents: list[str]) -> tuple[str, list[str]]:
    """Extract a public TikTok bio and its explicitly published profile link."""
    texts: list[str] = []
    links: list[str] = []

    def add_text(value: object) -> None:
        text = str(value or "").strip()
        if text and text not in texts:
            texts.append(text)

    def add_link(value: object) -> None:
        if isinstance(value, str):
            candidate = value.strip()
            if candidate.startswith(("http://", "https://")) and candidate not in links:
                links.append(candidate)
        elif isinstance(value, dict):
            for key in ("link", "url", "bio_url", "bioUrl"):
                add_link(value.get(key))

    def visit(value: object) -> None:
        if isinstance(value, dict):
            for key in ("signature", "bio_description", "bioDescription", "bio"):
                add_text(value.get(key))
            for key in ("bioLink", "bio_link", "bio_url", "bioUrl", "website"):
                add_link(value.get(key))
            for nested in value.values():
                visit(nested)
        elif isinstance(value, list):
            for nested in value:
                visit(nested)

    for document in documents:
        try:
            visit(json.loads(document))
        except (TypeError, ValueError):
            continue
    return "\n".join(texts), links


class SocialProfileCrawler:
    """Read public profile bios and explicitly published external websites."""

    def __init__(
        self,
        logger: logging.Logger,
        status: Callable[[str], None] | None = None,
        show_browser: bool = False,
        interval: float = 1.0,
        timeout_seconds: float = 30,
        scan_public_websites: bool = True,
    ):
        self.logger = logger
        self.status = status or (lambda _message: None)
        self.show_browser = show_browser
        self.interval = interval
        self.timeout_ms = int(timeout_seconds * 1000)
        self.scan_public_websites = scan_public_websites
        self.youtube = BrowserCrawler(
            logger, self.status, show_browser=show_browser, interval=interval,
            timeout_seconds=timeout_seconds,
        )

    def crawl_excel(
        self,
        input_file: Path,
        output_file: Path,
        sheet_name: str = "KOL List",
        platforms: set[str] | None = None,
        start_row: int = 2,
        end_row: int | None = None,
        limit: int | None = None,
        stop_event: threading.Event | None = None,
    ) -> int:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl，请执行：python -m pip install -r requirements.txt") from exc
        try:
            from playwright.sync_api import sync_playwright
        except ImportError as exc:
            raise RuntimeError("缺少 playwright，请执行：python -m pip install -r requirements.txt") from exc

        stop_event = stop_event or threading.Event()
        workbook = load_workbook(input_file)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"找不到工作表：{sheet_name}")
        sheet = workbook[sheet_name]
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        required = {"平台", "账号", "主页链接"}
        missing = required - headers.keys()
        if missing:
            raise ValueError(f"工作表缺少列：{'、'.join(sorted(missing))}")
        result_columns = self._ensure_result_columns(sheet, headers)
        wanted = {normalize_platform(x) for x in (platforms or SUPPORTED_PLATFORMS)}
        output_file.parent.mkdir(parents=True, exist_ok=True)
        processed = 0

        with sync_playwright() as playwright:
            browser = self.youtube._launch_browser(playwright)
            context = browser.new_context(
                locale="en-US",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
                ),
                viewport={"width": 1365, "height": 900},
            )
            page = context.new_page()
            page.set_default_timeout(self.timeout_ms)
            finder = PublicEmailFinder(
                status=self._emit, interval=min(self.interval, 0.5),
                timeout=min(self.timeout_ms / 1000, 15),
            )
            try:
                final_row = min(sheet.max_row, end_row) if end_row is not None else sheet.max_row
                for row_number in range(max(2, start_row), final_row + 1):
                    if stop_event.is_set() or (limit is not None and processed >= limit):
                        break
                    platform = normalize_platform(sheet.cell(row_number, headers["平台"]).value)
                    if platform not in wanted:
                        continue
                    account = sheet.cell(row_number, headers["账号"]).value
                    url = profile_url(
                        platform, account, sheet.cell(row_number, headers["主页链接"]).value
                    )
                    if platform not in SUPPORTED_PLATFORMS:
                        result = ProfileResult(crawl_status=f"不支持的平台：{platform}")
                    elif not url:
                        result = ProfileResult(crawl_status="缺少主页链接和账号")
                    else:
                        self._emit(f"第 {row_number} 行，读取 {platform}: {url}")
                        try:
                            result = self._crawl_profile(page, finder, platform, url)
                        except Exception as exc:  # keep the batch moving
                            self.logger.exception("主页采集失败：%s", url)
                            result = ProfileResult(crawl_status=f"失败：{type(exc).__name__}: {exc}")
                    self._write_result(sheet, row_number, result_columns, result)
                    processed += 1
                    if processed % 5 == 0:
                        workbook.save(output_file)
                    if self.interval:
                        time.sleep(self.interval)
            finally:
                workbook.save(output_file)
                context.close()
                browser.close()
        self._emit(f"已处理 {processed} 条，结果保存至：{output_file}")
        return processed

    def _crawl_profile(self, page, finder: PublicEmailFinder, platform: str, url: str) -> ProfileResult:
        if platform == "youtube":
            channel = self.youtube._read_channel(page, url)
            text = str(channel.get("description", ""))
            links = [str(x) for x in channel.get("links", [])]
            verification = bool(channel.get("email_verification_required"))
        else:
            text, links, page_status = self._read_generic_profile(page, platform, url)
            verification = page_status == "需要登录/验证"

        contacts = extract_public_contacts(text + "\n" + "\n".join(links))
        external = select_public_links(url, links)
        emails = [x.strip() for x in contacts["email"].split("|") if x.strip()]
        sources = {email: "公开主页简介" for email in emails}
        if self.scan_public_websites and external:
            discovery = finder.find(external)
            for email in discovery.emails:
                if email not in emails:
                    emails.append(email)
                sources[email] = discovery.sources.get(email, "公开官网")
        email_text = " | ".join(emails)
        status = classify_email_status(email_text, verification)
        source_text = " | ".join(dict.fromkeys(sources.values()))
        if verification and not email_text:
            crawl_status = "需要登录/验证"
        elif platform != "youtube" and page_status != "成功":
            crawl_status = page_status
        else:
            crawl_status = "成功"
        return ProfileResult(email_text, status, source_text, " | ".join(external), crawl_status)

    def _read_generic_profile(self, page, platform: str, url: str) -> tuple[str, list[str], str]:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=self.timeout_ms)
        except Exception as exc:
            # Some social pages keep background requests open; retain rendered content if available.
            if not page.url or page.url == "about:blank":
                raise exc
        page.wait_for_timeout(1800)
        data = page.evaluate(
            r"""platform => {
                const selectors = {
                    instagram: ['header'],
                    tiktok: ['[data-e2e="user-bio"]', '[data-e2e="user-link"]'],
                    x: ['[data-testid="UserDescription"]', '[data-testid="UserUrl"]']
                }[platform] || ['main'];
                const nodes = selectors.flatMap(selector => Array.from(document.querySelectorAll(selector)));
                const uniqueNodes = Array.from(new Set(nodes));
                return {
                    title: document.title || '',
                    description: document.querySelector('meta[name=description]')?.content || '',
                    ogDescription: document.querySelector('meta[property="og:description"]')?.content || '',
                    profileText: uniqueNodes.map(node => node.innerText || '').join('\n').slice(0, 50000),
                    pageText: document.body?.innerText?.slice(0, 50000) || '',
                    links: uniqueNodes.flatMap(node => [
                        ...(node.matches?.('a[href]') ? [node] : []),
                        ...Array.from(node.querySelectorAll('a[href]'))
                    ])
                        .map(a => a.href).filter(Boolean),
                    hydration: platform === 'tiktok'
                        ? Array.from(document.querySelectorAll(
                            'script#__UNIVERSAL_DATA_FOR_REHYDRATION__, script#SIGI_STATE, script[type="application/json"]'
                          )).map(script => script.textContent || '')
                          .filter(text => text.includes('userInfo') || text.includes('bioLink') || text.includes('signature'))
                          .map(text => text.slice(0, 2_000_000))
                        : []
                };
            }""",
            platform,
        )
        text = "\n".join(
            str(data.get(key, ""))
            for key in ("title", "description", "ogDescription", "profileText")
        )
        hydration_data = False
        if platform == "tiktok":
            bio_text, bio_links = extract_tiktok_profile_data(data.get("hydration", []))
            hydration_data = bool(bio_text or bio_links)
            if bio_text:
                text = "\n".join((text, bio_text))
            for bio_link in bio_links:
                if bio_link not in data["links"]:
                    data["links"].append(bio_link)
        lowered = (text + "\n" + str(data.get("pageText", ""))).casefold()
        for found_url in PUBLIC_URL_RE.findall(text):
            links = data.setdefault("links", [])
            cleaned = found_url.rstrip(".,;:!?)）]")
            if cleaned not in links:
                links.append(cleaned)
        gated = any(
            marker in lowered
            for marker in (
                "captcha", "unusual traffic", "verify you are human", "log in to continue",
                "login to continue", "登录以继续", "登入以繼續",
            )
        )
        has_profile_data = hydration_data or bool(str(data.get("profileText", "")).strip()) or any(
            value and str(value).strip() not in {"TikTok", "Instagram", "X"}
            for value in (data.get("description"), data.get("ogDescription"))
        )
        if platform == "tiktok" and not has_profile_data:
            oembed_text, oembed_status = self._read_tiktok_oembed(url)
            if oembed_text:
                text = "\n".join((text, oembed_text))
                has_profile_data = True
            elif oembed_status:
                return text, list(dict.fromkeys(data.get("links", []))), oembed_status
        if gated:
            page_status = "需要登录/验证"
        elif not has_profile_data:
            page_status = "页面未提供公开资料（可能受访问限制）"
        else:
            page_status = "成功"
        return text, list(dict.fromkeys(data.get("links", []))), page_status

    @staticmethod
    def _read_tiktok_oembed(profile_url_value: str) -> tuple[str, str]:
        """Use TikTok's documented public oEmbed endpoint as a non-browser fallback."""
        try:
            response = requests.get(
                "https://www.tiktok.com/oembed",
                params={"url": profile_url_value},
                headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
                timeout=15,
            )
        except requests.RequestException:
            return "", "TikTok oEmbed 请求失败"
        content_type = response.headers.get("Content-Type", "").casefold()
        if response.ok and "json" in content_type:
            try:
                payload = response.json()
            except ValueError:
                return "", "TikTok oEmbed 返回无效数据"
            return "\n".join(
                str(payload.get(key, "")) for key in ("title", "author_name", "author_url")
            ), ""
        if "/about" in urlparse(response.url).path.casefold():
            return "", "TikTok 当前网络被重定向到地区介绍页，无法读取公开简介"
        return "", f"TikTok oEmbed 不可用（HTTP {response.status_code}）"

    @staticmethod
    def _ensure_result_columns(sheet, headers: dict[str, int]) -> dict[str, int]:
        result = {}
        # Some supplied workbooks have formatting extending hundreds of blank columns.
        # Append beside the last real header, not beside the last styled cell.
        next_column = max(headers.values(), default=0) + 1
        for name in RESULT_COLUMNS:
            if name not in headers:
                sheet.cell(1, next_column, name)
                headers[name] = next_column
                next_column += 1
            result[name] = headers[name]
        return result

    @staticmethod
    def _write_result(sheet, row: int, columns: dict[str, int], result: ProfileResult) -> None:
        from datetime import datetime

        values = {
            "联系邮箱": result.email,
            "邮箱状态": result.email_status,
            "邮箱来源": result.email_source,
            "公开外链": result.public_links,
            "采集状态": result.crawl_status,
            "采集时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        for name, value in values.items():
            sheet.cell(row, columns[name], value)

    def _emit(self, message: str) -> None:
        self.logger.info(message)
        self.status(message)
