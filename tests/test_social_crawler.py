import unittest
from pathlib import Path
from unittest.mock import Mock

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from youtube_collector.social_crawler import (
    normalize_platform,
    profile_url,
    select_public_links,
    SocialProfileCrawler,
    extract_tiktok_profile_data,
    unwrap_external_url,
    PUBLIC_URL_RE,
)


class SocialCrawlerTests(unittest.TestCase):
    def test_platform_and_missing_url_are_normalized(self):
        self.assertEqual(normalize_platform("Twitter"), "x")
        self.assertEqual(normalize_platform("Twitter/X"), "x")
        self.assertEqual(normalize_platform("X/Twitter"), "x")
        self.assertEqual(normalize_platform("Twitter - X"), "x")
        self.assertEqual(profile_url("youtube", "Fireship", ""), "https://www.youtube.com/@Fireship")
        self.assertEqual(profile_url("tiktok", "@ai.zapo", ""), "https://www.tiktok.com/@ai.zapo")

    def test_instagram_redirect_is_unwrapped(self):
        self.assertEqual(
            unwrap_external_url("https://l.instagram.com/?u=https%3A%2F%2Fcreator.example%2Fcontact"),
            "https://creator.example/contact",
        )

    def test_public_url_in_profile_description_is_detected(self):
        self.assertEqual(
            PUBLIC_URL_RE.findall("Web: https://t.co/abc123\nWelcome"),
            ["https://t.co/abc123"],
        )

    def test_tiktok_hydration_extracts_bio_and_aggregation_link(self):
        bio, links = extract_tiktok_profile_data([
            '{"userInfo":{"user":{"signature":"Email: hello@example.com",'
            '"bioLink":{"link":"https://linktr.ee/example"}}}}'
        ])
        self.assertIn("hello@example.com", bio)
        self.assertEqual(links, ["https://linktr.ee/example"])

    def test_only_deliberate_external_links_are_selected(self):
        links = select_public_links(
            "https://www.instagram.com/example/",
            [
                "https://www.instagram.com/accounts/login/",
                "https://x.com/example",
                "https://creator.example/",
                "mailto:hello@creator.example",
                "https://creator.example/",
            ],
        )
        self.assertEqual(links, ["https://creator.example/", "mailto:hello@creator.example"])

    def test_result_columns_follow_real_headers_not_styled_blank_cells(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 1, "平台")
        sheet.cell(1, 16, "数据更新时间")
        sheet.cell(1, 498).fill = PatternFill("solid", fgColor="FFFF00")
        headers = {"平台": 1, "数据更新时间": 16}
        columns = SocialProfileCrawler._ensure_result_columns(sheet, headers)
        self.assertEqual(columns["联系邮箱"], 17)
        self.assertEqual(columns["采集时间"], 22)

    def test_rows_with_previous_collection_result_are_skipped(self):
        workbook = Workbook()
        sheet = workbook.active
        columns = {"采集状态": 5, "采集时间": 6}
        self.assertFalse(SocialProfileCrawler._row_already_processed(sheet, 2, columns))
        sheet.cell(2, columns["采集状态"], "成功")
        self.assertTrue(SocialProfileCrawler._row_already_processed(sheet, 2, columns))

    def test_locked_output_has_actionable_error(self):
        workbook = Mock()
        workbook.save.side_effect = PermissionError(13, "Permission denied")
        with self.assertRaisesRegex(PermissionError, "关闭 Excel/WPS"):
            SocialProfileCrawler._save_workbook(workbook, Path("result.xlsx"))

    def test_public_profile_email_skips_gated_email_button(self):
        crawler = SocialProfileCrawler(Mock(), scan_public_websites=True)
        crawler.youtube._read_channel = Mock(return_value={
            "description": "Business: public@creator.test",
            "links": [],
            "email_verification_required": True,
        })
        crawler.youtube._reveal_channel_email = Mock(return_value="hidden@creator.test")

        result = crawler._crawl_profile(Mock(), Mock(), "youtube", "https://youtube.com/@demo")

        self.assertEqual(result.email, "public@creator.test")
        crawler.youtube._reveal_channel_email.assert_not_called()

    def test_gated_email_is_fallback_when_public_sources_are_empty(self):
        crawler = SocialProfileCrawler(
            Mock(), scan_public_websites=True, reveal_gated_email=True
        )
        page = Mock()
        crawler.youtube._read_channel = Mock(return_value={
            "description": "No public contact",
            "links": [],
            "email_verification_required": True,
        })
        crawler.youtube._reveal_channel_email = Mock(return_value="hidden@creator.test")

        result = crawler._crawl_profile(page, Mock(), "youtube", "https://youtube.com/@demo")

        self.assertEqual(result.email, "hidden@creator.test")
        crawler.youtube._reveal_channel_email.assert_called_once_with(page)

    def test_gated_email_button_is_disabled_by_default(self):
        crawler = SocialProfileCrawler(Mock(), scan_public_websites=True)
        page = Mock()
        crawler.youtube._read_channel = Mock(return_value={
            "description": "No public contact",
            "links": [],
            "email_verification_required": True,
        })
        crawler.youtube._reveal_channel_email = Mock(return_value="hidden@creator.test")

        result = crawler._crawl_profile(page, Mock(), "youtube", "https://youtube.com/@demo")

        self.assertEqual(result.email, "")
        self.assertEqual(result.crawl_status, "需要登录/验证")
        crawler.youtube._reveal_channel_email.assert_not_called()


if __name__ == "__main__":
    unittest.main()
