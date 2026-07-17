"""导入"邮箱采集结果"格式 Excel（Richup_Mango / Pippit_100 / Dola_UK_150）到大数据库。

这三份是 22-23 列的"邮箱采集结果"格式，与 KOL-Find（28列）不同，需要列名映射。
全部内容导入 kol_candidate（大数据库），有邮箱的同时选入 kol + kol_email。

用法：
  python scripts/import_email_collection.py <xlsx> --preset richup --dry-run
  python scripts/import_email_collection.py <xlsx> --preset pippit --commit --batch "pippit-100"
  python scripts/import_email_collection.py <xlsx> --preset dola --commit --batch "dola-150"

--preset 可选 richup/pippit/dola；不传则按 sheet 名自动识别。
"""
from __future__ import annotations

import argparse
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Union

from openpyxl import load_workbook
from sqlalchemy import func

from db import SessionLocal
from models import Kol, KolCandidate, KolEmail

EMAIL_RE = re.compile(r"[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)

# ===== 三份文件的格式预设 =====
# 每份定义：sheet 名 + 画像列（拼进 remark）+ 主要推荐产品列名
PRESETS: dict[str, dict] = {
    "richup": {
        "sheet": "KOL List",
        "remark_columns": [],  # Richup 无画像列
        "has_fit_columns": True,  # 有 适配Dreamina/Pippit/Kimi/Dola 4列
    },
    "pippit": {
        "sheet": "Pippit达人名单",
        "remark_columns": ["达人类别", "社会身份/头衔备注", "推荐内容角度", "内容证据"],
        "has_fit_columns": False,
    },
    "dola": {
        "sheet": "Dola英国新增150人",
        "remark_columns": ["达人画像", "优先级", "Dola核心场景", "推荐内容角度", "内容证据"],
        "has_fit_columns": False,
    },
}

# 适配产品的 ✓/✅ 类符号（非空非"否"即视为适配）
FIT_TRUE_MARKS = {"✓", "✅", "√", "是", "true", "yes", "1", "y"}


def as_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v).strip()


def _trunc(v: str | None, max_len: int) -> str | None:
    """截断字符串到 max_len，避免超 varchar 宽度。None 透传。"""
    if v is None:
        return None
    return v[:max_len] if len(v) > max_len else v


def parse_int(v: Any) -> int | None:
    """解析粉丝数/浏览量。支持纯数字；K/M 后缀兜底。"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return max(0, int(v))
    s = str(v).lower().replace(",", "").replace(" ", "").strip()
    mult = 1
    if s.endswith("k"):
        s, mult = s[:-1], 1_000
    elif s.endswith("m"):
        s, mult = s[:-1], 1_000_000
    try:
        return max(0, int(float(s) * mult))
    except ValueError:
        return None


def parse_emails(raw: Any) -> list[str]:
    """拆分多邮箱（| , ; 分隔），小写去重保序。"""
    text = as_text(raw).replace("|", " ").replace(",", " ").replace(";", " ")
    seen: list[str] = []
    for m in EMAIL_RE.findall(text):
        e = m.strip(".,;:()[]<>\"'").lower()
        if "@" in e and e not in seen:
            seen.append(e)
    return seen


def parse_datetime(v: Any):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = as_text(v)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def platform_normalize(raw: str) -> str:
    low = raw.lower().strip()
    return {"youtube": "YouTube", "instagram": "Instagram",
            "tiktok": "TikTok", "x": "X", "twitter": "X", "twitter/x": "X"}.get(low, raw.strip())


def build_fit_product(row: dict[str, Any], headers: list[str], preset: dict) -> str | None:
    """从适配列拼出 fit_product 字符串。

    Richup 优先用已有的"适配产品"列（逗号拼好的）；否则从 适配X 列里取✓的。
    Pippit/Dola 只有单一 适配X 列，直接取产品名。
    """
    # Richup 已有"适配产品"列
    if "适配产品" in headers:
        val = as_text(row.get("适配产品"))
        if val:
            return val
    # 从 适配X 列拼
    products: list[str] = []
    for h in headers:
        if not h or "适配" not in str(h) or h == "适配产品":
            continue
        val = as_text(row.get(h)).lower()
        if val in FIT_TRUE_MARKS or (val and val not in {"❌", "否", "no", "false", "0", "×", "x"}):
            # 列名"适配Pippit"→ 产品名"Pippit"
            product = str(h).replace("适配", "").strip()
            if product:
                products.append(product)
    return "，".join(products) if products else None


def build_remark(row: dict[str, Any], preset: dict) -> str | None:
    """把画像列拼成 remark 文本（标签: 值 换行拼接）。"""
    parts: list[str] = []
    for col in preset.get("remark_columns", []):
        val = as_text(row.get(col))
        if val:
            parts.append(f"{col}: {val}")
    return "\n".join(parts) if parts else None


def map_row_to_candidate(row: dict[str, Any], headers: list[str], preset: dict, source_row: int) -> dict[str, Any]:
    """把一行 Excel 数据映射成 kol_candidate 字段 dict。

    所有 varchar 字段按列宽截断，避免 ``value too long`` 错误。
    列宽对应 kol_candidate model 定义。
    """
    platform = platform_normalize(as_text(row.get("平台")))
    account = as_text(row.get("账号"))
    return {
        "source_row": source_row,
        "platform": _trunc(platform, 20),
        "account": _trunc(account, 200),
        "profile_url": _trunc(as_text(row.get("主页链接")) or None, 2048),
        "related_youtube": None,
        "yt_about_source": None,
        "discovery_method": None,
        "fit_product": _trunc(build_fit_product(row, headers, preset), 100),
        "recommend_product": _trunc(as_text(row.get("主要推荐产品")) or None, 50),
        "hit_keyword_count": None,
        "keyword_note": None,
        "crawl_priority": None,
        "email_crawler": None,
        "email_source_url": None,
        "email_verify_status": "待爬取",
        "country_region": _trunc(as_text(row.get("国家/地区")) or None, 100),
        "language": _trunc(as_text(row.get("语言")) or None, 50),
        "account_type": None,
        "followers": parse_int(row.get("粉丝数")),
        "avg_views": parse_int(row.get("10天平均浏览量")),
        "review_status": "待复核",
        "remark": build_remark(row, preset),
        "contact_email": _trunc(as_text(row.get("联系邮箱")) or None, 500),
        "email_status": _trunc(as_text(row.get("邮箱状态")) or None, 50),
        "email_source": _trunc(as_text(row.get("邮箱来源")) or None, 100),
        "other_links": as_text(row.get("公开外链")) or None,  # text 类型，不截断
        "collect_status": _trunc(as_text(row.get("采集状态")) or None, 100),
        "collected_at": parse_datetime(row.get("采集时间")),
    }


def detect_preset(sheet_names: list[str], headers: list[str]) -> str:
    """按 sheet 名/列名自动识别格式。"""
    hset = set(headers)
    if "Dola核心场景" in hset or "达人画像" in hset:
        return "dola"
    if "达人类别" in hset or "社会身份/头衔备注" in hset:
        return "pippit"
    if "适配Dreamina" in hset or "适配产品" in hset:
        return "richup"
    raise ValueError(f"无法识别格式。sheet={sheet_names}, 列={headers[:8]}...")


def load_rows(xlsx_source: Union[Path, bytes, io.BytesIO], preset_name: str | None) -> tuple[list[dict], list[str], dict]:
    """读 Excel，返回 (rows, headers, preset)。"""
    if isinstance(xlsx_source, (bytes, bytearray)):
        xlsx_source = io.BytesIO(xlsx_source)
    wb = load_workbook(xlsx_source, read_only=True, data_only=True)

    preset_name = preset_name or detect_preset(wb.sheetnames, [])
    if preset_name not in PRESETS:
        raise ValueError(f"未知 preset: {preset_name}，可选: {list(PRESETS.keys())}")
    preset = PRESETS[preset_name]
    sheet_name = preset["sheet"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 缺少「{sheet_name}」表，现有: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers = [as_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # 校验共有列
    required = ["平台", "账号", "主页链接", "联系邮箱"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(f"Excel 缺少必要列: {missing}")

    rows: list[dict] = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not values or not any(v is not None for v in values):
            continue
        row = dict(zip(headers, values))
        mapped = map_row_to_candidate(row, headers, preset, source_row=idx - 1)
        if not mapped["platform"] or not mapped["account"]:
            continue
        rows.append(mapped)
    wb.close()
    return rows, headers, preset


def run_import(
    xlsx_source: Union[Path, bytes, io.BytesIO],
    preset_name: str | None,
    commit: bool,
    batch: str,
    db=None,
) -> dict:
    """执行导入。返回统计 dict。"""
    rows, headers, preset = load_rows(xlsx_source, preset_name)
    stats = {
        "preset": preset_name or "auto",
        "candidate_total": len(rows),
        "candidate_inserted": 0,
        "candidate_skipped_dup": 0,
        "emailable": 0,
        "kol_inserted": 0,
        "kol_skipped_dup": 0,
        "kol_email_inserted": 0,
    }

    own_session = db is None
    if own_session:
        db = SessionLocal()
    try:
        existing_kol_emails = {(e[0] or "").lower() for e in db.query(Kol.email).all()}
        # 预查现有 kol（含对象），用于已存在时的画像补全
        existing_kol_map: dict[str, Kol] = {}
        for k in db.query(Kol).all():
            if k.email:
                existing_kol_map[k.email.lower()] = k
        # 预查现有 candidate，含 followers/avg_views 状态，用于增量补全
        existing_candidates = {
            (c.platform, c.account): c
            for c in db.query(KolCandidate).all()
        }
        # 同时维护一个 set 便于快速判断（不持有对象的轻量副本）
        existing_pa = set(existing_candidates.keys())

        for item in rows:
            pa = (item["platform"], item["account"])
            # 1) kol_candidate
            if pa in existing_pa:
                stats["candidate_skipped_dup"] += 1
                # 增量补全：这三份有粉丝数/浏览量/画像，而 KOL-Find 没有。
                # 若已存在行的这些字段为空，用新数据补上（不覆盖非空值）。
                if commit:
                    cand = existing_candidates[pa]
                    backfill_fields = [
                        ("followers", item.get("followers")),
                        ("avg_views", item.get("avg_views")),
                        ("country_region", item.get("country_region")),
                        ("language", item.get("language")),
                        ("contact_email", item.get("contact_email")),
                        ("email_status", item.get("email_status")),
                        ("email_source", item.get("email_source")),
                        ("fit_product", item.get("fit_product")),
                        ("recommend_product", item.get("recommend_product")),
                        ("remark", item.get("remark")),
                    ]
                    for field, value in backfill_fields:
                        if value is not None and not getattr(cand, field, None):
                            setattr(cand, field, value)
                            stats["candidate_enriched"] = stats.get("candidate_enriched", 0) + 1
            else:
                if commit:
                    db.add(KolCandidate(import_batch=batch, **item))
                existing_pa.add(pa)
                stats["candidate_inserted"] += 1

            # 2) 有邮箱 → kol + kol_email
            emails = parse_emails(item.get("contact_email"))
            if not emails:
                continue
            stats["emailable"] += 1
            primary = emails[0]
            if primary in existing_kol_emails:
                stats["kol_skipped_dup"] += 1
                # 增量补全：这三份有粉丝数/平台/账号等，而已存在的 kol 多为 Snov
                # 纯邮箱导入（这些字段空）。补空字段，不覆盖非空值。
                if commit:
                    existing = existing_kol_map.get(primary)
                    if existing:
                        kol_fields = [
                            ("followers", item.get("followers")),
                            ("subscribers", item.get("followers")),
                            ("platform", item["platform"]),
                            ("social_handle", item["account"]),
                            ("profile_url", item.get("profile_url")),
                            ("channel_url", item.get("profile_url")),
                            ("country", item.get("country_region")),
                            ("niche", item.get("recommend_product")),
                            ("content_category", item.get("recommend_product")),
                        ]
                        for field, value in kol_fields:
                            if value:
                                cur = getattr(existing, field, None)
                                if not cur or (field in ("followers","subscribers") and (cur or 0) == 0):
                                    setattr(existing, field, value)
                                    stats["kol_enriched"] = stats.get("kol_enriched", 0) + 1
            else:
                if commit:
                    kol = Kol(
                        name=(item["account"] or "")[:200],
                        email=primary[:200],
                        platform=item["platform"][:50] or None,
                        social_handle=(item["account"] or "")[:200] or None,
                        profile_url=(item.get("profile_url") or "")[:500] or None,
                        channel_url=(item.get("profile_url") or "")[:500] or None,
                        country=(item.get("country_region") or "")[:50] or None,
                        niche=(item.get("recommend_product") or "")[:100] or None,
                        content_category=(item.get("recommend_product") or "")[:150] or None,
                        followers=item.get("followers") or 0,
                        subscribers=item.get("followers") or 0,
                        source=f"邮箱采集结果 | {batch}",
                        status="pending",
                    )
                    db.add(kol)
                    db.flush()
                    for i, em in enumerate(emails):
                        db.add(KolEmail(
                            kol_id=kol.id, email=em, email_normalized=em,
                            is_primary=(i == 0), source=f"邮箱采集结果 | {batch}",
                        ))
                existing_kol_emails.add(primary)
                stats["kol_inserted"] += 1
                stats["kol_email_inserted"] += len(emails)

        if commit:
            db.commit()
        else:
            db.rollback()
    except Exception:
        db.rollback()
        raise
    finally:
        if own_session:
            db.close()
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="导入邮箱采集结果 Excel（三份格式）到大数据库")
    parser.add_argument("source", type=Path)
    parser.add_argument("--preset", choices=list(PRESETS.keys()), default=None, help="格式预设，不传则自动识别")
    parser.add_argument("--commit", action="store_true")
    parser.add_argument("--batch", default=f"email-collection-{datetime.utcnow().strftime('%Y%m%d')}")
    args = parser.parse_args()

    mode = "写库" if args.commit else "预览(dry-run)"
    print(f"导入邮箱采集结果 [{mode}] preset={args.preset or 'auto'} batch={args.batch}")
    stats = run_import(args.source, args.preset, args.commit, args.batch)
    print("\n=== 结果 ===")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
