"""Clean a KOL collection workbook and optionally merge it into the database."""
from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import func

from db import SessionLocal
from models import Kol, KolEmail, ProjectAssessment
from services.email_utils import normalize_email


EMAIL_RE = re.compile(r"[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
SOURCE_LABEL = "Dola UK KOL Expansion 150 - 邮箱采集结果"
SOURCE_MARKER = f"[{SOURCE_LABEL}]"


def infer_project_code(label: str) -> str | None:
    """从批次来源标签推断 project_code。Dola/Pippit 是并列项目。

    "Dola UK KOL Expansion ..." -> "dola_uk"
    "Pippit ..."                -> "pippit_2026"
    无法识别 -> None（评估仍写入，但 project_code 为空，便于人工归类）
    """
    low = (label or "").lower()
    if "dola" in low:
        return "dola_uk"
    if "pippit" in low:
        return "pippit_2026"
    return None


# 当前批次对应的项目代码（clean_row/classify_and_merge 共用）。
PROJECT_CODE = infer_project_code(SOURCE_LABEL)

SOURCE_COLUMNS = [
    "平台", "账号", "昵称", "主页链接", "粉丝数", "10天平均浏览量", "国家/地区", "语言",
    "达人画像", "优先级", "内容赛道", "适配Dola", "Dola核心场景", "推荐内容角度",
    "内容证据", "来源链接", "数据更新时间", "联系邮箱", "邮箱状态", "邮箱来源",
    "公开外链", "采集状态", "采集时间",
]

CLEAN_COLUMNS = [
    "name", "email", "platform", "social_handle", "profile_url", "followers", "country",
    "priority", "content_category", "company_site", "recent_videos", "source", "contact_notes",
    "source_row", "import_action",
]


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def normalize_email(raw: Any) -> tuple[str, list[str]]:
    matches: list[str] = []
    # EMAIL_RE 的字符类含 ``|``，会把 ``a@x.com|b@y.com`` 的 ``|b@y.com`` 当合法邮箱
    # 匹配（``|`` 留在 local part）。导入前先把分隔符 | 换成空格，强制断词。
    text = as_text(raw).replace("mailto:", "").replace("|", " ")
    for match in EMAIL_RE.findall(text):
        candidate = match.strip(".,;:()[]<>\"'").lower()
        local, _, domain = candidate.partition("@")
        if not local or not domain or ".." in candidate or domain.startswith("-") or domain.endswith("-"):
            continue
        if candidate not in matches:
            matches.append(candidate)
    return (matches[0] if matches else ""), matches[1:]


def normalize_url(raw: Any) -> str:
    value = as_text(raw)
    if not value:
        return ""
    if value.startswith("www."):
        value = "https://" + value
    try:
        parts = urlsplit(value)
        if parts.scheme.lower() not in {"http", "https"} or not parts.netloc:
            return ""
        return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), parts.path, parts.query, ""))
    except ValueError:
        return ""


def parse_count(raw: Any) -> int:
    if raw is None or raw == "":
        return 0
    if isinstance(raw, (int, float)):
        return max(0, int(raw))
    value = as_text(raw).lower().replace(",", "").replace(" ", "")
    multiplier = 1
    if value.endswith("k"):
        value, multiplier = value[:-1], 1_000
    elif value.endswith("m"):
        value, multiplier = value[:-1], 1_000_000
    try:
        return max(0, int(float(value) * multiplier))
    except ValueError:
        return 0


def stronger_priority(left: str, right: str) -> str:
    candidates = [priority for priority in (left, right) if priority in PRIORITY_ORDER]
    return min(candidates, key=PRIORITY_ORDER.get) if candidates else ""


def make_notes(row: dict[str, Any], alternate_emails: list[str]) -> str:
    details = [
        ("达人画像", row.get("达人画像")),
        ("适配 Dola", row.get("适配Dola")),
        ("Dola 核心场景", row.get("Dola核心场景")),
        ("推荐内容角度", row.get("推荐内容角度")),
        ("内容证据", row.get("内容证据")),
        ("语言", row.get("语言")),
        ("10天平均浏览量", row.get("10天平均浏览量")),
        ("邮箱状态", row.get("邮箱状态")),
        ("邮箱来源", row.get("邮箱来源")),
        ("采集状态", row.get("采集状态")),
        ("采集时间", row.get("采集时间")),
        ("来源链接", row.get("来源链接")),
        ("备用邮箱", ", ".join(alternate_emails)),
    ]
    return "\n".join(f"{label}: {as_text(value)}" for label, value in details if as_text(value))


def clean_row(row: dict[str, Any], source_row: int) -> tuple[dict[str, Any] | None, str]:
    email, alternates = normalize_email(row.get("联系邮箱"))
    if not email:
        return None, "缺少有效联系邮箱"
    handle = as_text(row.get("账号"))
    name = as_text(row.get("昵称")) or handle.lstrip("@") or email.split("@", 1)[0]
    if not name:
        return None, "缺少达人名称"
    priority = as_text(row.get("优先级")).upper()
    if priority not in PRIORITY_ORDER:
        priority = ""
    evidence = as_text(row.get("内容证据"))
    profile_url = normalize_url(row.get("主页链接")) or normalize_url(row.get("来源链接"))

    # ===== Tier 1 标准化结构化字段（直接从 Excel 列读，不再只塞 contact_notes）=====
    # 语言转 BCP47，复用迁移逻辑保证与历史回填一致。
    from migrations.data_migration import _to_avg_views, _to_lang, _parse_collect_at

    # 适配/核心场景列名随项目变化：Dola 批次是 "适配Dola"/"Dola核心场景"，
    # Pippit 批次是 "适配Pippit"/"Pippit核心场景"。用 .get() 容错任一缺失。
    fit_raw = as_text(row.get("适配Dola") or row.get("适配Pippit"))
    fit_status = "fit" if fit_raw.strip() in {"✓", "√"} else None
    core_scenario = as_text(row.get("Dola核心场景") or row.get("Pippit核心场景"))[:200] or None
    collect_at = _parse_collect_at(as_text(row.get("采集时间")))

    return {
        "name": name[:200],
        "email": email[:200],
        "platform": as_text(row.get("平台"))[:50],
        "social_handle": handle[:200],
        "profile_url": profile_url[:500],
        "followers": parse_count(row.get("粉丝数")),
        "country": as_text(row.get("国家/地区"))[:50],
        "priority": priority,
        "content_category": as_text(row.get("内容赛道"))[:150],
        "company_site": normalize_url(row.get("公开外链"))[:500],
        "recent_videos": [evidence] if evidence else [],
        "source": SOURCE_LABEL,
        "contact_notes": make_notes(row, alternates),
        "source_row": source_row,
        "import_action": "",
        # 结构化新字段
        "avg_views_10d": _to_avg_views(as_text(row.get("10天平均浏览量"))),
        "language": _to_lang(as_text(row.get("语言"))),
        "email_status": as_text(row.get("邮箱状态"))[:50] or None,
        "email_source": as_text(row.get("邮箱来源"))[:100] or None,
        "collect_status": as_text(row.get("采集状态"))[:50] or None,
        "collect_at": collect_at,
        "source_link": normalize_url(row.get("来源链接"))[:2048] or None,
        "fit_project_code": PROJECT_CODE,
        "project_code": PROJECT_CODE,
        "fit_status": fit_status,
        "core_scenario": core_scenario,
        "recommend_angle": as_text(row.get("推荐内容角度")) or None,
        "kol_category": as_text(row.get("达人画像"))[:100] or None,
        "alternate_emails": alternates,
    }, ""


def merge_duplicate(base: dict[str, Any], incoming: dict[str, Any]) -> None:
    for key in ("platform", "social_handle", "profile_url", "country", "content_category", "company_site"):
        if not base.get(key) and incoming.get(key):
            base[key] = incoming[key]
    base["followers"] = max(base.get("followers", 0), incoming.get("followers", 0))
    base["priority"] = stronger_priority(base.get("priority", ""), incoming.get("priority", ""))
    base["recent_videos"] = list(dict.fromkeys(base.get("recent_videos", []) + incoming.get("recent_videos", [])))
    if incoming.get("contact_notes") and incoming["contact_notes"] not in base.get("contact_notes", ""):
        base["contact_notes"] = (base.get("contact_notes", "") + "\n\n" + incoming["contact_notes"]).strip()


def load_source(path: Path) -> tuple[list[list[Any]], list[dict[str, Any]], list[dict[str, Any]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [as_text(cell.value) for cell in sheet[1]]
    missing = [column for column in SOURCE_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"源表缺少字段: {', '.join(missing)}")

    raw_rows: list[list[Any]] = [headers]
    cleaned_by_email: dict[str, dict[str, Any]] = {}
    rejected: list[dict[str, Any]] = []
    source_duplicates = 0
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_rows.append(list(values))
        row = dict(zip(headers, values))
        cleaned, reason = clean_row(row, row_number)
        if not cleaned:
            rejected.append({
                "source_row": row_number,
                "name": as_text(row.get("昵称")),
                "raw_email": as_text(row.get("联系邮箱")),
                "reason": reason,
            })
            continue
        if cleaned["email"] in cleaned_by_email:
            merge_duplicate(cleaned_by_email[cleaned["email"]], cleaned)
            source_duplicates += 1
        else:
            cleaned_by_email[cleaned["email"]] = cleaned
    workbook.close()
    return raw_rows, list(cleaned_by_email.values()), rejected, source_duplicates


def _write_assessment_and_emails(db, kol: Kol, item: dict[str, Any]) -> None:
    """upsert project_assessment 与 kol_email。

    - project_assessment: 同 (project_code, kol_id) 先删后插，实现最新批次覆盖评估。
    - kol_email: 主邮箱 + 备用邮箱幂等追加（已存在则跳过）。
    必须在 kol.id 已就绪后调用（新增分支先 db.flush()）。
    """
    project_code = item.get("project_code")
    if project_code:
        # 删除该 KOL 在本项目下的旧评估，再插入新的（最新批次覆盖）。
        db.query(ProjectAssessment).filter(
            ProjectAssessment.project_code == project_code,
            ProjectAssessment.kol_id == kol.id,
        ).delete(synchronize_session=False)
        if item.get("fit_status") or item.get("core_scenario") or \
           item.get("recommend_angle") or item.get("kol_category"):
            db.add(ProjectAssessment(
                project_code=project_code,
                kol_id=kol.id,
                fit_status=item.get("fit_status"),
                core_scenario=item.get("core_scenario"),
                recommend_angle=item.get("recommend_angle"),
                kol_category=item.get("kol_category"),
                collected_at=item.get("collect_at"),
            ))

    # kol_email：主邮箱（来自 kol.email）+ 备用邮箱。
    primary = normalize_email(kol.email)
    emails_to_add = []
    if primary:
        emails_to_add.append((primary, True))
    for alt in item.get("alternate_emails") or []:
        alt_norm = normalize_email(alt)
        if alt_norm and alt_norm != primary:
            emails_to_add.append((alt_norm, False))
    # 幂等：跳过已存在的 (kol_id, email_normalized)。
    existing_norms = {e.email_normalized for e in
                      db.query(KolEmail).filter(KolEmail.kol_id == kol.id).all()}
    for email_norm, is_primary in emails_to_add:
        if email_norm in existing_norms:
            continue
        db.add(KolEmail(
            kol_id=kol.id,
            email=email_norm,
            email_normalized=email_norm,
            is_primary=is_primary,
            source=SOURCE_LABEL,
        ))


def classify_and_merge(cleaned: list[dict[str, Any]], commit: bool) -> dict[str, int]:
    db = SessionLocal()
    counts = {"new": 0, "enriched": 0, "unchanged": 0}
    try:
        for item in cleaned:
            existing = db.query(Kol).filter(func.lower(Kol.email) == item["email"]).first()
            if not existing:
                item["import_action"] = "新增"
                counts["new"] += 1
                if commit:
                    kol = Kol(
                        name=item["name"], full_name=item["name"], email=item["email"],
                        channel_url=item["profile_url"] or None, profile_url=item["profile_url"] or None,
                        platform=item["platform"] or None, social_handle=item["social_handle"] or None,
                        followers=item["followers"], subscribers=item["followers"],
                        country=item["country"] or None, priority=item["priority"] or None,
                        content_category=item["content_category"] or None,
                        niche=item["content_category"] or None, company_site=item["company_site"] or None,
                        recent_videos=item["recent_videos"] or None, source=item["source"],
                        contact_notes=(f"{SOURCE_MARKER}\n{item['contact_notes']}" if item["contact_notes"] else SOURCE_MARKER),
                        status="pending",
                        # Tier 1 结构化新列
                        avg_views_10d=item.get("avg_views_10d"),
                        language=item.get("language"),
                        email_status=item.get("email_status"),
                        email_source=item.get("email_source"),
                        collect_status=item.get("collect_status"),
                        collect_at=item.get("collect_at"),
                        source_link=item.get("source_link"),
                        fit_project_code=item.get("fit_project_code"),
                    )
                    db.add(kol)
                    db.flush()  # 拿到 kol.id，供子表外键用
                    _write_assessment_and_emails(db, kol, item)
                continue

            changed = False
            fields = {
                "full_name": item["name"], "platform": item["platform"],
                "social_handle": item["social_handle"], "profile_url": item["profile_url"],
                "channel_url": item["profile_url"], "country": item["country"],
                "content_category": item["content_category"], "niche": item["content_category"],
                "company_site": item["company_site"],
            }
            for field, value in fields.items():
                if value and not getattr(existing, field, None):
                    if commit:
                        setattr(existing, field, value)
                    changed = True
            # 粉丝数：最新批次覆盖（CONSTRAINTS §4.3），而非取 max。
            # 历史值若需保留，走 metric_snapshot 表（Tier 2+）；当前 Tier 1 单值即最新。
            new_followers = item["followers"]
            current_followers = max(existing.followers or 0, existing.subscribers or 0)
            if new_followers != current_followers:
                if commit:
                    existing.followers = new_followers
                    existing.subscribers = new_followers
                changed = True
            priority = stronger_priority(existing.priority or "", item["priority"])
            if priority and priority != (existing.priority or ""):
                if commit:
                    existing.priority = priority
                changed = True
            videos = list(dict.fromkeys((existing.recent_videos or []) + item["recent_videos"]))
            if videos != (existing.recent_videos or []):
                if commit:
                    existing.recent_videos = videos
                changed = True

            # Tier 1 结构化新列：最新批次覆盖（与 followers 同策略，CONSTRAINTS §4.3）。
            new_fields = {
                "avg_views_10d": item.get("avg_views_10d"),
                "language": item.get("language"),
                "email_status": item.get("email_status"),
                "email_source": item.get("email_source"),
                "collect_status": item.get("collect_status"),
                "collect_at": item.get("collect_at"),
                "source_link": item.get("source_link"),
                "fit_project_code": item.get("fit_project_code"),
            }
            for field, value in new_fields.items():
                if value is not None and getattr(existing, field, None) != value:
                    if commit:
                        setattr(existing, field, value)
                    changed = True
            # project_assessment 与 kol_email：upsert（同 (project_code,kol_id) 覆盖评估，
            # 新邮箱追加）。补充分支也调，确保重复导入不漏评估。
            if commit and item.get("project_code"):
                _write_assessment_and_emails(db, existing, item)
            if item["contact_notes"] and SOURCE_MARKER not in (existing.contact_notes or ""):
                if commit:
                    existing.contact_notes = (
                        (existing.contact_notes or "") + f"\n\n{SOURCE_MARKER}\n" + item["contact_notes"]
                    ).strip()
                changed = True
            if SOURCE_LABEL not in (existing.source or ""):
                if commit:
                    existing.source = " | ".join(filter(None, [existing.source, SOURCE_LABEL]))[:200]
                changed = True
            item["import_action"] = "补充已有记录" if changed else "数据库已存在"
            counts["enriched" if changed else "unchanged"] += 1
        if commit:
            db.commit()
        else:
            db.rollback()
        return counts
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def style_table(sheet, table_name: str) -> None:
    if sheet.max_row < 1 or sheet.max_column < 1:
        return
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F2")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    if sheet.max_row >= 2:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False
        )
        sheet.add_table(table)


def write_output(
    path: Path,
    raw_rows: list[list[Any]],
    cleaned: list[dict[str, Any]],
    rejected: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "导入摘要"
    summary_sheet.append(["指标", "结果"])
    for key, value in summary.items():
        summary_sheet.append([key, value])
    style_table(summary_sheet, "ImportSummaryTable")
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 58

    clean_sheet = workbook.create_sheet("清洗后 KOL")
    clean_sheet.append(CLEAN_COLUMNS)
    for item in cleaned:
        clean_sheet.append([
            " | ".join(item[column]) if isinstance(item[column], list) else item[column]
            for column in CLEAN_COLUMNS
        ])
    style_table(clean_sheet, "CleanedKOLTable")
    widths = [24, 32, 14, 22, 42, 12, 12, 10, 18, 40, 42, 28, 72, 12, 18]
    for index, width in enumerate(widths, start=1):
        clean_sheet.column_dimensions[clean_sheet.cell(1, index).column_letter].width = width

    rejected_sheet = workbook.create_sheet("拒绝记录")
    rejected_sheet.append(["source_row", "name", "raw_email", "reason"])
    for item in rejected:
        rejected_sheet.append([item["source_row"], item["name"], item["raw_email"], item["reason"]])
    style_table(rejected_sheet, "RejectedKOLTable")
    for column, width in zip("ABCD", [12, 28, 38, 30]):
        rejected_sheet.column_dimensions[column].width = width

    raw_sheet = workbook.create_sheet("原始数据")
    for row in raw_rows:
        raw_sheet.append(row)
    style_table(raw_sheet, "RawSourceTable")
    for column in raw_sheet.columns:
        letter = column[0].column_letter
        longest = max(len(as_text(cell.value)) for cell in column[:40])
        raw_sheet.column_dimensions[letter].width = min(max(longest + 2, 12), 42)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--commit", action="store_true")
    args = parser.parse_args()

    raw_rows, cleaned, rejected, source_duplicates = load_source(args.source)
    database_counts = classify_and_merge(cleaned, args.commit)
    summary = {
        "源文件": args.source.name,
        "源数据行数": len(raw_rows) - 1,
        "有效唯一邮箱": len(cleaned),
        "源表重复邮箱": source_duplicates,
        "拒绝记录": len(rejected),
        "数据库新增": database_counts["new"],
        "补充已有记录": database_counts["enriched"],
        "数据库已存在且无需更新": database_counts["unchanged"],
        "执行模式": "已导入数据库" if args.commit else "预检查（未写数据库）",
        "清洗规则": "邮箱小写并校验；按邮箱去重；粉丝数转整数；优先级限制 P0-P3；URL 规范化；原始数据完整保留",
    }
    write_output(args.output, raw_rows, cleaned, rejected, summary)
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
