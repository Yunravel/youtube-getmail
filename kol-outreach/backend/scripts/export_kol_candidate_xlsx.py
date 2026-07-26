"""从 kol_candidate 表导出 KOL 清单为 xlsx（按产品分 sheet）。

用法::

    # 导出 Hypic + SCRL 两项目合计清单（默认）
    python -m backend.scripts.export_kol_candidate_xlsx --output ./hypic_scrl_kol.xlsx

    # 只导出某个产品
    python -m backend.scripts.export_kol_candidate_xlsx --output ./hypic.xlsx --product Hypic

    # 加阈值过滤：仅保留 ≥5000 粉、近 10 天均播 ≥1000 的行
    python -m backend.scripts.export_kol_candidate_xlsx --output out.xlsx \\
        --min-followers 5000 --min-avg-views 1000

筛选逻辑：按 fit_product 字段做子串匹配（fit_product 是逗号分隔多值串，
如 "Hypic，SCRL"）。fit_product 为空或匹配不到的行不进入任何产品 sheet，
但会进入"未归类"sheet 供人工排查。

不写库，纯查询 + 导出。字段映射对齐 kol-find 默认中文表头。
"""
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from db import SessionLocal
from models import KolCandidate

# kol-find 默认中文表头（与 references/config-template.md 的 output_schema 一致）
# (中文列名, KolCandidate 属性名, 是否参与阈值过滤)
COLUMNS: list[tuple[str, str, bool]] = [
    ("平台", "platform", False),
    ("账号", "account", False),
    ("昵称", "_display_name", False),  # 候选池无独立昵称列，用 account 兜底
    ("主页链接", "profile_url", False),
    ("粉丝数", "followers", True),
    ("10天平均浏览量", "avg_views", True),
    ("国家/地区", "country_region", False),
    ("语言", "language", False),
    ("内容赛道", "account_type", False),  # account_type 存"个人创作者/工作室"等，近似赛道
    ("适配产品", "fit_product", False),
    ("主要推荐产品", "recommend_product", False),
    ("数据更新时间", "collected_at", False),
    ("联系方式类型", "email_source", False),
    ("联系方式", "contact_email", False),
]

# 默认导出的两个项目
DEFAULT_PRODUCTS = ["Hypic", "SCRL"]

# 阈值默认：0 表示不过滤（导出全部）。kol-find 默认建议 YT/TT/IG ≥5000 粉、
# 近 10 天均播 ≥1000-2000，但爬取入库时不过滤，导出时可按需裁剪。
DEFAULT_MIN_FOLLOWERS = 0
DEFAULT_MIN_AVG_VIEWS = 0


def _attr(row: KolCandidate, attr: str) -> Any:
    """安全取属性；_display_name 特殊处理（候选池无独立昵称列，用 account 兜底）。"""
    if attr == "_display_name":
        return getattr(row, "account", "") or ""
    return getattr(row, attr, None)


def _fmt_value(value: Any) -> Any:
    """时间转 isoformat 字符串，None→空串，其余原样（让 openpyxl 处理数字）。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return value


def query_candidates(
    db: Session,
    products: list[str],
    min_followers: int,
    min_avg_views: int,
) -> dict[str, list[KolCandidate]]:
    """按产品分组查询候选行。

    返回 {product: [rows]}；另含 key "_unclassified" 存不属于任何指定产品的行。
    同一行若 fit 多个产品，会出现在多个分组里（符合 kol-find 的"一人可适多产品"）。

    阈值过滤：followers/avg_views 为 NULL 时不被阈值排除（数据缺失不应误删），
    只有"有值且低于阈值"才排除。
    """
    q = db.query(KolCandidate)
    if min_followers > 0:
        # NULL 保留：KolCandidate.followers IS NULL OR >= min
        q = q.filter(
            (KolCandidate.followers.is_(None)) | (KolCandidate.followers >= min_followers)
        )
    if min_avg_views > 0:
        q = q.filter(
            (KolCandidate.avg_views.is_(None)) | (KolCandidate.avg_views >= min_avg_views)
        )
    all_rows = q.order_by(KolCandidate.followers.desc().nullslast()).all()

    grouped: dict[str, list[KolCandidate]] = {p: [] for p in products}
    grouped["_unclassified"] = []
    for row in all_rows:
        fit = (row.fit_product or "").replace("，", ",").replace("、", ",")
        fit_set = {f.strip() for f in fit.split(",") if f.strip()}
        matched = [p for p in products if p in fit_set]
        if matched:
            for p in matched:
                grouped[p].append(row)
        else:
            grouped["_unclassified"].append(row)
    return grouped


def style_table(sheet) -> None:
    """复用 import_kol_xlsx 的蓝表头样式。"""
    if sheet.max_row < 1 or sheet.max_column < 1:
        return
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F2")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    # 列宽：前几列窄，profile_url/contact_notes 宽
    widths = [10, 22, 20, 40, 10, 12, 12, 8, 18, 16, 14, 18, 14, 30]
    for idx, w in enumerate(widths, start=1):
        col = sheet.cell(1, idx).column_letter
        sheet.column_dimensions[col].width = w


def _write_product_sheet(wb: Workbook, title: str, rows: list[KolCandidate]) -> None:
    sheet = wb.create_sheet(title[:31])  # Excel sheet 名限 31 字符
    headers = [c[0] for c in COLUMNS]
    sheet.append(headers)
    for row in rows:
        sheet.append([_fmt_value(_attr(row, c[1])) for c in COLUMNS])
    if sheet.max_row >= 2:
        table = Table(displayName=f"T{title}"[:120].replace(" ", "_"), ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True,
            showFirstColumn=False, showLastColumn=False,
        )
        sheet.add_table(table)
    style_table(sheet)


def write_workbook(
    path: Path,
    grouped: dict[str, list[KolCandidate]],
    products: list[str],
    min_followers: int,
    min_avg_views: int,
) -> None:
    wb = Workbook()
    # 摘要 sheet
    summary = wb.active
    summary.title = "导出摘要"
    summary.append(["指标", "结果"])
    summary.append(["导出时间", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    summary.append(["筛选产品", "、".join(products)])
    summary.append(["最低粉丝阈值", min_followers or "不过滤"])
    summary.append(["最低10天均播阈值", min_avg_views or "不过滤"])
    total = 0
    for p in products:
        n = len(grouped.get(p, []))
        summary.append([f"{p} 候选数", n])
        total += n
    summary.append(["合计（含跨产品重复）", total])
    summary.append(["未归类行数", len(grouped.get("_unclassified", []))])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 50
    style_table(summary)

    # 各产品 sheet
    for p in products:
        _write_product_sheet(wb, p, grouped.get(p, []))

    # 未归类（供人工排查，可能因 fit_product 拼写差异漏归类）
    if grouped.get("_unclassified"):
        _write_product_sheet(wb, "未归类", grouped["_unclassified"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="导出 kol_candidate 为按产品分 sheet 的 xlsx")
    parser.add_argument("--output", required=True, type=Path, help="输出 xlsx 路径")
    parser.add_argument(
        "--product", action="append", default=None,
        help=f"筛选产品（可多次传；默认 {DEFAULT_PRODUCTS}）",
    )
    parser.add_argument("--min-followers", type=int, default=DEFAULT_MIN_FOLLOWERS)
    parser.add_argument("--min-avg-views", type=int, default=DEFAULT_MIN_AVG_VIEWS)
    args = parser.parse_args()

    products = args.product or DEFAULT_PRODUCTS
    db = SessionLocal()
    try:
        grouped = query_candidates(
            db, products, args.min_followers, args.min_avg_views
        )
        write_workbook(
            args.output, grouped, products,
            args.min_followers, args.min_avg_views,
        )
    finally:
        db.close()

    # 打印摘要到 stdout
    print(f"已导出: {args.output}")
    for p in products:
        print(f"  {p}: {len(grouped.get(p, []))} 行")
    print(f"  未归类: {len(grouped.get('_unclassified', []))} 行")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
