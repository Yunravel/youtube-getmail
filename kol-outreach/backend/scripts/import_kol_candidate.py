"""导入 KOL-Find 多平台候选池到 kol_candidate 表，并把有邮箱的选入 kol + kol_email 表。

用法：
  python scripts/import_kol_candidate.py <xlsx路径> --dry-run
  python scripts/import_kol_candidate.py <xlsx路径> --commit --batch "kol-find-20260716"

数据流：
  Excel「全部候选」5103 行
    → kol_candidate（全量，UNIQUE(platform, account) 去重）
    → 有 contact_email 的行 → kol 表（邮箱去重，已存在跳过）
                            → kol_email 表（多邮箱拆分入库）

幂等：重复跑只插新行，不产生重复。
"""
from __future__ import annotations

import argparse
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Union

from openpyxl import load_workbook

from db import SessionLocal
from models import Kol, KolCandidate, KolEmail
from services.email_utils import normalize_email
# 解析工具抽取到共享模块，避免与 import_email_collection.py / 采集器三处分叉。
from scripts._parse_utils import (
    EMAIL_RE,
    as_text,
    parse_datetime,
    parse_emails,
    parse_int,
    platform_normalize,
)

# Excel 列名 → kol_candidate 字段名
COLUMN_MAP = {
    "序号": "source_row",
    "平台": "platform",
    "账号": "account",
    "主页链接": "profile_url",
    "关联YouTube账号": "related_youtube",
    "YouTube About来源": "yt_about_source",
    "发现方式": "discovery_method",
    "适配产品": "fit_product",
    "主要推荐产品": "recommend_product",
    "命中检索词数": "hit_keyword_count",
    "发现关键词/种子说明": "keyword_note",
    "抓取优先级": "crawl_priority",
    "邮箱（爬虫回填）": "email_crawler",
    "邮箱来源URL": "email_source_url",
    "邮箱核验状态": "email_verify_status",
    "国家/地区": "country_region",
    "语言": "language",
    "账号类型": "account_type",
    "粉丝数": "followers",
    "近期平均播放": "avg_views",
    "人工复核状态": "review_status",
    "备注": "remark",
    "联系邮箱": "contact_email",
    "邮箱状态": "email_status",
    "邮箱来源": "email_source",
    "其他链接": "other_links",
    "采集状态": "collect_status",
    "采集时间": "collected_at",
}

def load_rows(xlsx_source: Union[Path, bytes, io.BytesIO]) -> list[dict[str, Any]]:
    """读 Excel「全部候选」表，返回字段 dict 列表。

    xlsx_source 可以是文件路径（Path/str）、字节流（bytes）或 BytesIO，
    供 CLI 和 HTTP 上传接口共用。
    """
    # openpyxl 对裸 bytes 不直接支持，统一包成 BytesIO。
    if isinstance(xlsx_source, (bytes, bytearray)):
        xlsx_source = io.BytesIO(xlsx_source)
    wb = load_workbook(xlsx_source, read_only=True, data_only=True)
    if "全部候选" not in wb.sheetnames:
        raise ValueError(f"Excel 缺少「全部候选」表，现有: {wb.sheetnames}")
    ws = wb["全部候选"]
    headers = [as_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # 校验列名
    missing = [cn for cn in COLUMN_MAP if cn not in headers]
    if missing:
        raise ValueError(f"Excel 缺少列: {missing}")

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values or not any(v is not None for v in values):
            continue
        row = dict(zip(headers, values))
        item: dict[str, Any] = {}
        for cn, fn in COLUMN_MAP.items():
            val = row.get(cn)
            if fn in ("source_row", "hit_keyword_count"):
                item[fn] = parse_int(val)
            elif fn in ("followers", "avg_views"):
                item[fn] = parse_int(val)
            elif fn == "collected_at":
                item[fn] = parse_datetime(val)
            elif fn == "platform":
                item[fn] = platform_normalize(as_text(val))
            else:
                item[fn] = as_text(val) or None
        if not item.get("platform") or not item.get("account"):
            continue
        rows.append(item)
    wb.close()
    return rows


def upsert_candidates(
    rows: list[dict[str, Any]],
    db,
    batch: str,
    *,
    source_label: str = "KOL-Find 候选池",
    commit: bool = True,
    backfill_existing: bool = True,
) -> dict:
    """把已规范化的 candidate 行 dict 列表双写入库。

    这是 Excel 导入与采集器共用的入库核心。rows 里每个 dict 的键名须为
    kol_candidate 的 ORM 字段名（platform/account/profile_url/contact_email/...），
    非中文 Excel 列名。

    - kol_candidate：按 (platform, account) 去重；新行直接写，已存在行可选回填空字段
    - 有 contact_email 的行 → 晋升 kol（按主邮箱去重）+ kol_email（全部邮箱）

    backfill_existing=True（默认）：已存在的 candidate 行，若其某字段为空而新数据有值，
    则回填（不覆盖非空值）。这对采集器至关重要——历史 Excel 导入留下的 contact_email=None
    行，采集到邮箱后必须能更新进去。回填字段：邮箱相关 + 粉丝数 + 国家 + 画像等。

    幂等：重复跑只插新行/补空字段，不覆盖已有值，不产生重复。
    返回统计 dict。
    """
    stats = {
        "candidate_total": len(rows),
        "candidate_inserted": 0,
        "candidate_skipped_dup": 0,
        "candidate_enriched": 0,
        "emailable": 0,
        "kol_inserted": 0,
        "kol_skipped_dup": 0,
        "kol_email_inserted": 0,
    }

    # 预查现有 kol 邮箱（小写），避免逐行查询
    existing_kol_emails = {
        (e[0] or "").lower() for e in db.query(Kol.email).all()
    }
    # 预查现有 candidate (platform, account) → 对象映射，用于回填
    existing_candidates: dict[tuple, KolCandidate] = {}
    for c in db.query(KolCandidate).all():
        existing_candidates[(c.platform, c.account)] = c

    # 已存在行需要回填的字段（仅当旧值为空时写入新值）
    backfill_fields = [
        "contact_email", "email_crawler", "email_status",
        "email_verify_status", "email_source", "email_source_url",
        "collect_status", "collected_at",
        "followers", "avg_views", "country_region", "language",
        "account_type", "fit_product", "recommend_product",
        "profile_url", "related_youtube", "yt_about_source",
        "discovery_method", "keyword_note", "hit_keyword_count",
        "remark", "other_links",
    ]

    for item in rows:
        pa = (item["platform"], item["account"])
        # 1) kol_candidate（去重 platform+account）
        if pa in existing_candidates:
            stats["candidate_skipped_dup"] += 1
            # 增量回填空字段（采集器核心价值：给历史空邮箱行补邮箱）
            if backfill_existing and commit:
                cand = existing_candidates[pa]
                for field in backfill_fields:
                    new_val = item.get(field)
                    if new_val is not None and not getattr(cand, field, None):
                        setattr(cand, field, new_val)
                        stats["candidate_enriched"] += 1
        else:
            if commit:
                db.add(KolCandidate(import_batch=batch, **item))
            existing_candidates[pa] = item  # 占位，避免同批次重复
            stats["candidate_inserted"] += 1

        # 2) 有邮箱 → 选入 kol + kol_email
        emails = parse_emails(item.get("contact_email"))
        if not emails:
            continue
        stats["emailable"] += 1
        primary_email = emails[0]

        if primary_email in existing_kol_emails:
            stats["kol_skipped_dup"] += 1
        else:
            if commit:
                kol = Kol(
                    name=item["account"][:200],
                    email=primary_email[:200],
                    platform=item["platform"][:50] or None,
                    social_handle=item["account"][:200] or None,
                    profile_url=(item.get("profile_url") or "")[:500] or None,
                    channel_url=(item.get("profile_url") or "")[:500] or None,
                    country=(item.get("country_region") or "")[:50] or None,
                    niche=(item.get("recommend_product") or "")[:100] or None,
                    content_category=(item.get("recommend_product") or "")[:150] or None,
                    # 粉丝数（followers/subscribers 同源；subscribers 是 YouTube 订阅数）
                    followers=item.get("followers") or 0,
                    subscribers=item.get("followers") or 0,
                    # 采集扩展字段（Tier 1 标准化列）
                    language=(item.get("language") or "")[:50] or None,
                    email_status=(item.get("email_status") or "")[:50] or None,
                    email_source=(item.get("email_source") or "")[:100] or None,
                    collect_status=(item.get("collect_status") or "")[:100] or None,
                    source_link=(item.get("yt_about_source") or "")[:2048] or None,
                    fit_project_code=(item.get("recommend_product") or "")[:100] or None,
                    source=f"{source_label} | {batch}",
                    status="pending",
                )
                db.add(kol)
                db.flush()  # 拿 kol.id
                # kol_email：全部邮箱入库
                for i, em in enumerate(emails):
                    db.add(KolEmail(
                        kol_id=kol.id,
                        email=em,
                        email_normalized=normalize_email(em),
                        is_primary=(i == 0),
                        source=f"{source_label} | {batch}",
                    ))
            existing_kol_emails.add(primary_email)
            stats["kol_inserted"] += 1
            stats["kol_email_inserted"] += len(emails)

    if commit:
        db.commit()
    else:
        db.rollback()
    return stats


def run_import(
    xlsx_source: Union[Path, bytes, io.BytesIO],
    commit: bool,
    batch: str,
    db=None,
) -> dict:
    """执行 Excel 导入。返回统计 dict。

    db 为 None 时内部创建独立 session（CLI 用）；传入时复用（HTTP 接口用，
    以便请求结束统一提交/回滚）。
    """
    rows = load_rows(xlsx_source)

    own_session = db is None
    if own_session:
        db = SessionLocal()
    try:
        return upsert_candidates(rows, db, batch, commit=commit)
    except Exception:
        db.rollback()
        raise
    finally:
        if own_session:
            db.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="导入 KOL-Find 候选池")
    parser.add_argument("source", type=Path, help="xlsx 文件路径")
    parser.add_argument("--commit", action="store_true", help="实际写库（默认 dry-run）")
    parser.add_argument("--batch", type=str, default=f"kol-find-{datetime.utcnow().strftime('%Y%m%d')}")
    args = parser.parse_args()

    mode = "写库" if args.commit else "预览(dry-run)"
    print(f"导入候选池 [{mode}] batch={args.batch}")
    print(f"读取: {args.source}")
    stats = run_import(args.source, args.commit, args.batch)
    print("\n=== 结果 ===")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
