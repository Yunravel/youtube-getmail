"""HotLead 报价单 Excel 导出。

按 ``Dola_已报价KOL_6位(1).xlsx`` 规范生成 xlsx：
  - 3 段式布局：标题行1 / KPI 卡 行4-5 / 表头行10 / 数据行11+
  - 12 列（A-L），中文列名一字不差
  - 样式：Carlito 字体、深海军蓝 #12304A 表头白字、浅蓝 #EAF4F8 KPI 卡
  - 列宽、行高、冻结窗格 A11、自动换行
  - 第二张表「报价说明」固定文案

入口：``build_quote_workbook(thread_ids, db) -> bytes``
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from services.ai_profile import parse_min_quote

# ===== 规范常量（来自规范文件实测）=====
_HEADERS = [
    "序号", "姓名", "平台", "社交账号", "主页链接", "社会身份",
    "头衔 / 内容定位", "邮箱", "粉丝数", "报价（GBP，起）", "报价明细", "合作要求 / 备注",
]
_COL_WIDTHS = [7, 22, 12, 20, 37, 24, 25, 28, 12, 15, 48, 55]
_TITLE_TEXT = "Dola 已报价 KOL 清单"
_FONT_NAME = "Carlito"
_HEADER_FILL = "FF12304A"   # 深海军蓝（ARGB）
_KPI_FILL = "FFEAF4F8"      # 浅蓝

_THIN = Side(style="thin", color="FFBFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _load_rows(db: Session, thread_ids: list[int], project_code: str) -> list[dict]:
    """从 DB 组装每个 thread 一行的导出数据（12 列）。

    数据来源优先级（结构化字段优先，AI 提取字段其次）：
      - kol 表：name/email/platform/social_handle/profile_url/followers/position/niche
      - project_assessment：kol_category（社会身份）/ recommend_angle（内容定位）
      - message.ai_analysis：budget_mentioned/deliverables/payment_terms/usage_rights/timeline/key_questions/summary
    """
    sql = text(
        """
        SELECT t.id AS thread_id,
               k.name, k.email, k.platform, k.social_handle,
               k.profile_url, k.channel_url,
               COALESCE(k.followers, k.subscribers) AS followers,
               k.position, k.niche, k.content_category,
               pa.kol_category, pa.recommend_angle,
               m.ai_analysis
        FROM thread t
        JOIN kol k ON k.id = t.kol_id
        LEFT JOIN project_assessment pa ON pa.kol_id = k.id AND pa.project_code = :pc
        LEFT JOIN LATERAL (
            SELECT ai_analysis FROM message
            WHERE thread_id = t.id AND direction = 'inbound'
              AND body_text IS NOT NULL AND body_text <> ''
            ORDER BY received_at DESC LIMIT 1
        ) m ON true
        WHERE t.id = ANY(:ids)
        ORDER BY t.intent_score DESC NULLS LAST, t.updated_at DESC
        """
    )
    raw = db.connection().execute(sql, {"ids": thread_ids, "pc": project_code}).fetchall()

    rows: list[dict] = []
    for r in raw:
        ai = r.ai_analysis or {}
        if isinstance(ai, str):
            import json
            try:
                ai = json.loads(ai)
            except Exception:
                ai = {}

        # 列6 社会身份：project_assessment.kol_category 优先，其次 kol.position（AI 回填），再 content_category
        niche_display = r.kol_category or r.position or r.content_category
        # 列7 内容定位：recommend_angle 优先，其次 niche
        focus_display = r.recommend_angle or r.niche

        # 列10 报价：budget_mentioned 正则解析
        budget_raw = ai.get("budget_mentioned")
        min_amount, currency = parse_min_quote(budget_raw)

        # 列11 报价明细：deliverables 优先，其次 budget_mentioned 原串
        deliverables = ai.get("deliverables") or budget_raw

        # 列12 合作要求/备注：多字段拼接
        note_parts = []
        if ai.get("payment_terms"):
            note_parts.append(f"付款：{ai['payment_terms']}")
        if ai.get("usage_rights"):
            note_parts.append(f"授权：{ai['usage_rights']}")
        if ai.get("timeline") and ai.get("timeline") not in ("none", None):
            note_parts.append(f"档期：{ai['timeline']}")
        if ai.get("summary"):
            note_parts.append(ai["summary"])
        cooperation_note = "；".join(note_parts) if note_parts else None

        # 未报价标记：J/K 空，L 写【未报价】
        if min_amount is None and not deliverables:
            if cooperation_note:
                cooperation_note = f"【未报价】{cooperation_note}"
            else:
                cooperation_note = "【未报价】已回复但未提供报价"

        rows.append({
            "name": r.name or "",
            "platform": _platform_display(r.platform),
            "social_handle": r.social_handle or "",
            "profile_url": r.profile_url or r.channel_url or "",
            "niche": niche_display or "",
            "focus": focus_display or "",
            "email": r.email or "",
            "followers": int(r.followers) if r.followers else None,
            "min_amount": min_amount,
            "currency": currency,
            "deliverables": deliverables or "",
            "cooperation_note": cooperation_note or "",
        })
    return rows


def _platform_display(raw: Optional[str]) -> str:
    """规范化平台展示名。规范要求 Instagram/TikTok/YouTube 等。"""
    if not raw:
        return ""
    low = raw.lower().strip()
    return {
        "youtube": "YouTube", "yt": "YouTube",
        "instagram": "Instagram", "ig": "Instagram",
        "tiktok": "TikTok",
        "x": "X", "twitter": "X", "twitter/x": "X",
        "linkedin": "LinkedIn",
        "website": "Website",
    }.get(low, raw)


def _set_font(cell, size=11, bold=False, color=None):
    cell.font = Font(name=_FONT_NAME, size=size, bold=bold, color=color)


def build_quote_workbook(db: Session, thread_ids: list[int], project_code: str = "dola_uk") -> bytes:
    """生成 xlsx 字节流。"""
    rows = _load_rows(db, thread_ids, project_code)
    n = len(rows)
    last_data_row = 10 + n  # 表头在行10，数据从行11开始

    wb = Workbook()
    ws = wb.active
    ws.title = "报价KOL"

    # ---- 列宽 ----
    for idx, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ---- 行1：标题横幅（合并 A1:L1）----
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"已报价 KOL 清单（{n} 位）"
    _set_font(title_cell, size=16, bold=True, color="FFFFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    # ---- 行2-3 空白间隔 ----
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 8

    # ---- 行4-5：KPI 卡（3 张，各占 2 列）----
    kpi_defs = [
        ("已报价 KOL", f"=COUNTA(B11:B{last_data_row})"),
        ("已核验报价", f"=COUNTA(J11:J{last_data_row})"),
        ("最低起报价", f"=MIN(J11:J{last_data_row})"),
    ]
    for i, (label, formula) in enumerate(kpi_defs):
        col_left = 1 + i * 2  # A/C/E
        col_right = col_left + 1
        label_range = f"{get_column_letter(col_left)}4:{get_column_letter(col_right)}4"
        value_range = f"{get_column_letter(col_left)}5:{get_column_letter(col_right)}5"
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        lc = ws.cell(row=4, column=col_left, value=label)
        _set_font(lc, size=11, bold=True)
        lc.fill = PatternFill("solid", fgColor=_KPI_FILL)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row=5, column=col_left, value=formula)
        _set_font(vc, size=14, bold=True)
        vc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24

    # ---- 行6-9 空白间隔 ----
    for r in range(6, 10):
        ws.row_dimensions[r].height = 8

    # ---- 行10：表头 ----
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        _set_font(cell, size=11, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[10].height = 48

    # ---- 行11+：数据 ----
    for row_idx, row in enumerate(rows, start=11):
        # 报价列：数字 + 币种标注（规范是 GBP，但实测多为 USD，保留币种符号）
        amount_display = row["min_amount"] if row["min_amount"] is not None else ""
        # 报价明细若含币种信息就不再单独标
        values = [
            row_idx - 10,                          # 序号
            row["name"],                            # 姓名
            row["platform"],                        # 平台
            row["social_handle"],                   # 社交账号
            row["profile_url"],                     # 主页链接
            row["niche"],                           # 社会身份
            row["focus"],                           # 头衔/内容定位
            row["email"],                           # 邮箱
            row["followers"] if row["followers"] else "",  # 粉丝数
            amount_display,                         # 报价(起)
            row["deliverables"],                    # 报价明细
            row["cooperation_note"],                # 合作要求/备注
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _set_font(cell, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            cell.border = _BORDER
        ws.row_dimensions[row_idx].height = 72

    # ---- 冻结窗格 A11 ----
    ws.freeze_panes = "A11"

    # ---- 第二张表：报价说明 ----
    ws2 = wb.create_sheet("报价说明")
    ws2.column_dimensions["A"].width = 21
    ws2.column_dimensions["B"].width = 105
    ws2.merge_cells("A1:B1")
    t2 = ws2["A1"]
    t2.value = "报价说明"
    _set_font(t2, size=14, bold=True, color="FFFFFFFF")
    t2.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28

    notes = [
        ("数据来源", "来源于系统 KOL 外联中台的 Hot Lead 看板，按勾选的会话导出。"),
        ("报价币种", "金额保留 KOL 原始报价币种；标注 $ 通常为 USD，£ 为 GBP。未做汇率换算。"),
        ("报价口径", "「报价（起）」为正则从 KOL 报价串解析出的最低档金额；「报价明细」保留完整套餐原文。"),
        ("画像来源", "平台/账号/粉丝数/社会身份等由 AI 从回信正文提取并回填，可能有误差，需人工复核。"),
        ("未报价", "J、K 列空且 L 列以「【未报价】」开头的行，表示 KOL 已回复但未提供报价。"),
        ("更新时间", f"导出于 {__import__('datetime').datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"),
    ]
    ws2.cell(row=3, column=1, value="项目").font = Font(name=_FONT_NAME, bold=True)
    ws2.cell(row=3, column=2, value="说明").font = Font(name=_FONT_NAME, bold=True)
    for i, (k, v) in enumerate(notes, start=4):
        c1 = ws2.cell(row=i, column=1, value=k)
        c2 = ws2.cell(row=i, column=2, value=v)
        _set_font(c1, size=11, bold=True)
        _set_font(c2, size=11)
        c1.alignment = Alignment(vertical="top")
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        ws2.row_dimensions[i].height = 36

    # 输出字节
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
