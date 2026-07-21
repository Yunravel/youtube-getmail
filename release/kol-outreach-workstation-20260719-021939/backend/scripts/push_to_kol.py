"""把爬虫产出的 Excel 推送到 kol-outreach 大数据库。

爬虫跑完后调用本脚本，自动判断 Excel 格式并上传到对应导入接口，
不再需要手动打开浏览器上传。

用法：
  python push_to_kol.py <xlsx路径>                      # 推到生产（默认）
  python push_to_kol.py <xlsx路径> --url http://localhost:8000  # 推到本地后端
  python push_to_kol.py <xlsx路径> --token xxx          # 带 token（接口启用鉴权时）

格式自动判断（无需手动指定）：
  - 含「全部候选」表（28列）→ /api/kols/import-candidate（KOL-Find 候选池）
  - 含 Richup/Pippit/Dola 特征列   → /api/kols/import-email-collection（邮箱采集结果）

环境变量（可选，命令行参数优先）：
  KOL_API_URL     默认 https://kol.20020525.xyz
  KOL_IMPORT_TOKEN  接口 token（接口启用鉴权时需要）
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from urllib.parse import urljoin

import openpyxl

DEFAULT_URL = "https://kol.20020525.xyz"


def detect_format(xlsx_path: Path) -> str:
    """判断 Excel 格式，返回 'candidate' 或 'email-collection'。

    candidate: KOL-Find 多平台候选池（含「全部候选」表，28 列）
    email-collection: 邮箱采集结果（Richup/Pippit/Dola，22-23 列）

    遍历所有 sheet 找数据表（Richup 第一个 sheet 是「需求总结」，数据在「KOL List」）。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    # 优先看有无「全部候选」表（KOL-Find 候选池）
    if "全部候选" in sheet_names:
        wb.close()
        return "candidate"
    # 遍历所有 sheet，找邮箱采集格式的特征列
    required = {"平台", "账号", "主页链接", "联系邮箱"}
    for sn in sheet_names:
        ws = wb[sn]
        try:
            headers = {str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))}
        except StopIteration:
            continue
        if required <= headers:
            wb.close()
            return "email-collection"
    wb.close()
    raise ValueError(
        f"无法识别 Excel 格式。sheets={sheet_names}。"
        f"期望：含「全部候选」表（28列）或邮箱采集结果格式（含 平台/账号/主页链接/联系邮箱 列）。"
    )


def push(xlsx_path: Path, base_url: str, token: str | None, max_retries: int = 3) -> dict:
    """上传 xlsx 到对应接口，返回统计 dict。失败重试 max_retries 次。"""
    fmt = detect_format(xlsx_path)
    endpoint = "/api/kols/import-candidate" if fmt == "candidate" else "/api/kols/import-email-collection"
    url = urljoin(base_url + "/", endpoint.lstrip("/"))
    print(f"格式: {fmt} → 接口: {endpoint}")

    headers = {}
    if token:
        headers["X-Import-Token"] = token

    data = xlsx_path.read_bytes()
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            import requests
            resp = requests.post(
                url,
                files={"file": (xlsx_path.name, data,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                headers=headers,
                timeout=300,
            )
            if resp.status_code == 200:
                return resp.json()
            # 401/403 不重试（鉴权问题，重试无用）
            if resp.status_code in (401, 403):
                print(f"鉴权失败 ({resp.status_code}): {resp.text[:200]}")
                print("提示：接口启用了 token 鉴权，请用 --token 或环境变量 KOL_IMPORT_TOKEN 传入正确 token。")
                sys.exit(1)
            # 422 格式错误不重试
            if resp.status_code == 422:
                print(f"格式错误 (422): {resp.text[:300]}")
                sys.exit(1)
            last_err = f"HTTP {resp.status_code}: {resp.text[:200]}"
        except Exception as e:
            last_err = str(e)
        print(f"  第 {attempt}/{max_retries} 次失败: {last_err}")
        if attempt < max_retries:
            time.sleep(2 * attempt)
    raise RuntimeError(f"推送失败（重试 {max_retries} 次）: {last_err}")


def main() -> int:
    parser = argparse.ArgumentParser(description="推送爬虫产出的 Excel 到 kol-outreach 大数据库")
    parser.add_argument("source", type=Path, help="xlsx 文件路径")
    parser.add_argument("--url", default=os.environ.get("KOL_API_URL", DEFAULT_URL),
                        help=f"目标 API 地址（默认 {DEFAULT_URL}）")
    parser.add_argument("--token", default=os.environ.get("KOL_IMPORT_TOKEN"),
                        help="接口 token（环境变量 KOL_IMPORT_TOKEN 优先）")
    args = parser.parse_args()

    if not args.source.exists():
        print(f"文件不存在: {args.source}")
        return 1

    print(f"推送: {args.source.name} ({args.source.stat().st_size // 1024}KB) → {args.url}")
    try:
        stats = push(args.source, args.url, args.token)
    except (RuntimeError, ValueError) as e:
        print(f"\n✗ {e}")
        return 1

    print("\n=== 推送成功 ===")
    if "candidate_inserted" in stats:
        print(f"  候选库: 新增 {stats.get('candidate_inserted', 0)} / 跳过 {stats.get('candidate_skipped_dup', 0)}"
              f"（共 {stats.get('candidate_total', 0)}）")
        print(f"  选入发信池: 新增 {stats.get('kol_inserted', 0)} / 已存在 {stats.get('kol_skipped_dup', 0)}"
              f"（有邮箱 {stats.get('emailable', 0)}）")
        if stats.get("kol_email_inserted"):
            print(f"  邮箱记录: 新增 {stats['kol_email_inserted']}")
        if stats.get("candidate_enriched") or stats.get("kol_enriched"):
            print(f"  补全字段: candidate={stats.get('candidate_enriched', 0)} kol={stats.get('kol_enriched', 0)}")
    else:
        for k, v in stats.items():
            print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
